"""
T-1 mobility workbook importer.
Source: "T1 Measured CG ... .xlsx" (sensitive -- NOT committed to git).

Reads the vehicle CG inputs, stored SFs / axle loads, the shelter component
table and the reverse-CG payload from the two CG variants:
  - "measured" : inputs on 'T1 Measured CG'; SFs on 'M.Stability Assess (Measured)'.
  - "theory"   : inputs on the summary row of 'M.Stability Assessment (Theory)';
                 SFs on the same theory sheet.

Unlike the E2 importer (.xls via xlrd) the T1 file is .xlsx, so this module
uses openpyxl (data_only=True -- reads the last-saved cached formula results).
Grouped axle limits are supplied by mobility_profiles.T1_PROFILE.make_vehicle,
so the validated two-support-line engine runs unchanged (axles 1+2 -> front
group, 3+4 -> rear group).

Path resolution: T1_MOBILITY_XLSX environment variable, else the known local
Downloads path. workbook_path() returns None when neither exists, letting the
workbook-backed tests skip cleanly.
"""
import os
import warnings
from dataclasses import dataclass
from typing import Optional

from mobility_engine import Vehicle
from mobility_profiles import T1_PROFILE
from cg_engine import CGState, derive_payload_cg

WB_ENV = "T1_MOBILITY_XLSX"
WB_DEFAULT = r"C:\Users\qujia\Downloads\T1 Measured CG_23-10-2025 (11-3-2026) updated 1 (1).xlsx"


# ---------------------------------------------------------------------------
# Sheet names and cell references -- edit here if the workbook layout shifts.
# openpyxl uses 1-indexed A1 notation.
# ---------------------------------------------------------------------------

_SHEET_MEAS_CG = "T1 Measured CG"                 # measured laden vehicle inputs
_SHEET_UNLADEN = "Unladen Trsp N1"                # bare transporter (reverse-CG base)
_SHEET_THEORY = "M.Stability Assessment (Theory)"  # theory SFs + summary CG row
_SHEET_SF_MEAS = "M.Stability Assess (Measured)"   # measured SFs / axle loads
_SHEET_COMPONENTS = "Spinel T1 (updated)"          # shelter component mass table

# Measured / unladen CG cells (col D): D8 GW, D9 Xcg, D10 Ycg, D11 Zcg
_CG_CELLS = {"GW": "D8", "Xcg": "D9", "Ycg": "D10", "Zcg": "D11"}

# Theory summary row 6: C6 GW, D6 Xcg, E6 Ycg, F6 Zcg
_THEORY_CELLS = {"GW": "C6", "Xcg": "D6", "Ycg": "E6", "Zcg": "F6"}

# Stored SFs (same cell layout on measured + theory analysis sheets)
_STORED_SF = {
    "asc_60":  "AA17",
    "desc_60": "AA40",
    "kerb_30": "AK17",
    "road_30": "AK40",
    "corner":  "BH77",
}

# Stored axle loads: front P11, rear P12
_STORED_AXLE = {"front_kg": "P11", "rear_kg": "P12"}

# Component table on 'Spinel T1 (updated)', rows 4..78, columns:
#   A item | B description | C qty | D unit mass | E total mass | F X | G Y | H Z
_COMP_ROWS = range(4, 79)
_COMP_COLS = {"item": 1, "desc": 2, "qty": 3, "unit": 4, "total": 5, "x": 6, "y": 7, "z": 8}


@dataclass(frozen=True)
class T1ShelterComponent:
    """One row of the T1 shelter mass table (vehicle datum: X longitudinal,
    Y right-positive from centreline, Z from ground)."""
    item_no: int
    description: str
    qty: float
    unit_mass_kg: float
    total_mass_kg: float
    x_mm: float
    y_mm: float
    z_mm: float


# ---------------------------------------------------------------------------
# Path + open helpers
# ---------------------------------------------------------------------------

def workbook_path(path: Optional[str] = None) -> Optional[str]:
    """Resolve the workbook path (arg -> env -> default). Returns None if the
    resolved file does not exist, so callers/tests can skip rather than fail."""
    p = path or os.environ.get(WB_ENV) or WB_DEFAULT
    return p if p and os.path.exists(p) else None


def _open(path: Optional[str] = None):
    import openpyxl
    p = workbook_path(path)
    if p is None:
        raise FileNotFoundError(
            f"T1 workbook not found; set {WB_ENV} or place it at {WB_DEFAULT}")
    with warnings.catch_warnings():
        # openpyxl warns on the workbook's DrawingML extensions -- harmless here.
        warnings.simplefilter("ignore")
        return openpyxl.load_workbook(p, data_only=True)


def _cells(sheet, mapping: dict) -> dict:
    return {k: sheet[cell].value for k, cell in mapping.items()}


# ---------------------------------------------------------------------------
# Vehicles
# ---------------------------------------------------------------------------

def vehicle_measured_t1(path: Optional[str] = None) -> Vehicle:
    """Measured laden T1 vehicle from 'T1 Measured CG' (D8..D11)."""
    wb = _open(path)
    c = _cells(wb[_SHEET_MEAS_CG], _CG_CELLS)
    return T1_PROFILE.make_vehicle(
        "Spinel T1 (Measured CG)",
        gw_kg=float(c["GW"]), xcg_mm=float(c["Xcg"]),
        ycg_mm=float(c["Ycg"]), zcg_mm=float(c["Zcg"]))


def vehicle_theory_t1(path: Optional[str] = None) -> Vehicle:
    """Theory-buildup T1 vehicle from the summary row of the theory sheet (C6..F6)."""
    wb = _open(path)
    c = _cells(wb[_SHEET_THEORY], _THEORY_CELLS)
    return T1_PROFILE.make_vehicle(
        "Spinel T1 (Theory CG)",
        gw_kg=float(c["GW"]), xcg_mm=float(c["Xcg"]),
        ycg_mm=float(c["Ycg"]), zcg_mm=float(c["Zcg"]))


def unladen_cg_t1(path: Optional[str] = None) -> CGState:
    """Bare transporter CG from 'Unladen Trsp N1' (reverse-CG base)."""
    wb = _open(path)
    c = _cells(wb[_SHEET_UNLADEN], _CG_CELLS)
    return CGState("Unladen Trsp N1", float(c["GW"]),
                   float(c["Xcg"]), float(c["Ycg"]), float(c["Zcg"]))


# ---------------------------------------------------------------------------
# Stored values (for validation)
# ---------------------------------------------------------------------------

def _sf_sheet(variant: str) -> str:
    return _SHEET_SF_MEAS if variant == "measured" else _SHEET_THEORY


def stored_sf_map_t1(path: Optional[str] = None, variant: str = "measured") -> dict:
    """{label: stored_SF} from the measured or theory analysis sheet."""
    wb = _open(path)
    s = wb[_sf_sheet(variant)]
    return {k: float(s[cell].value) for k, cell in _STORED_SF.items()}


def stored_axle_map_t1(path: Optional[str] = None, variant: str = "measured") -> dict:
    """{front_kg, rear_kg} stored group loads from the analysis sheet."""
    wb = _open(path)
    s = wb[_sf_sheet(variant)]
    return {k: float(s[cell].value) for k, cell in _STORED_AXLE.items()}


# ---------------------------------------------------------------------------
# Component table + reverse-CG payload
# ---------------------------------------------------------------------------

def shelter_components_t1(path: Optional[str] = None) -> list:
    """Read the T1 shelter component table ('Spinel T1 (updated)', rows 4..78).
    Rows with a non-numeric / non-positive total mass are skipped."""
    wb = _open(path)
    s = wb[_SHEET_COMPONENTS]
    col = _COMP_COLS
    out = []
    for r in _COMP_ROWS:
        total = s.cell(row=r, column=col["total"]).value
        if not isinstance(total, (int, float)) or total <= 0:
            continue
        coords = [s.cell(row=r, column=col[k]).value for k in ("x", "y", "z")]
        if any(not isinstance(v, (int, float)) for v in coords):
            continue
        item = s.cell(row=r, column=col["item"]).value
        qty = s.cell(row=r, column=col["qty"]).value
        unit = s.cell(row=r, column=col["unit"]).value
        out.append(T1ShelterComponent(
            item_no=int(item) if isinstance(item, (int, float)) else 0,
            description=str(s.cell(row=r, column=col["desc"]).value or "").strip(),
            qty=float(qty) if isinstance(qty, (int, float)) else 0.0,
            unit_mass_kg=float(unit) if isinstance(unit, (int, float)) else 0.0,
            total_mass_kg=float(total),
            x_mm=float(coords[0]), y_mm=float(coords[1]), z_mm=float(coords[2]),
        ))
    return out


def payload_cg_from_measured_t1(path: Optional[str] = None) -> CGState:
    """Reverse-CG the shelter/payload: measured laden minus unladen transporter
    (the 'Re-engrg Laden Shelter Weight' method), via cg_engine.derive_payload_cg."""
    vm = vehicle_measured_t1(path)
    combined = CGState("T1 Measured (laden)", vm.gw_kg, vm.xcg_mm, vm.ycg_mm, vm.zcg_mm)
    base = unladen_cg_t1(path)
    return derive_payload_cg(combined, base, payload_name="T1 shelter/payload")


if __name__ == "__main__":
    if workbook_path() is None:
        raise SystemExit(f"T1 workbook not found; set {WB_ENV}")
    for variant, fn in (("MEASURED", vehicle_measured_t1), ("THEORY", vehicle_theory_t1)):
        v = fn()
        print(f"\n=== {variant} CG ===")
        print(f"  GW={v.gw_kg:.1f}  Xcg={v.xcg_mm:.2f}  Ycg={v.ycg_mm:.3f}  Zcg={v.zcg_mm:.3f}")
        print(f"  Stored SFs : {stored_sf_map_t1(variant=variant.lower())}")
        print(f"  Stored axle: {stored_axle_map_t1(variant=variant.lower())}")
    comps = shelter_components_t1()
    print(f"\nComponents: {len(comps)} rows, sum = {sum(c.total_mass_kg for c in comps):.1f} kg")
    pl = payload_cg_from_measured_t1()
    print(f"Payload rev-CG: {pl.mass_kg:.1f} kg  X={pl.x_mm:.3f}  Y={pl.y_mm:.3f}  Z={pl.z_mm:.3f}")
