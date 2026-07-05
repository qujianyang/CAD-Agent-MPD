"""
Proof harness: twist_lock_engine vs the T-1 workbook.
Run:  .\\mpd\\Scripts\\python.exe scripts\\validate_t1_twistlock.py

Reads sheets 'StrengthChk_TwistLock(Measured)' and '(Theory)' from the T-1
workbook, recomputes design force + per-lock force from the sheet's own
shelter weight via twist_lock_engine, and diffs the stored cells:
    F3 col F = Design Limit Load (N)      -> design_force_N
    F3 col H = Force acted on each lock   -> force_per_lock_N
Rows 6/7/8 = Longitudinal / Lateral / Vertical.
Expected: 12/12 (2 sheets x 3 axes x 2 columns... reported as 6 axis-checks
each covering both force columns -> 12 numeric comparisons).

Path override: T1_MOBILITY_XLSX environment variable.
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import warnings
warnings.filterwarnings("ignore")
import openpyxl

from twist_lock_engine import analyze_twist_locks

WB_DEFAULT = r"C:\Users\qujia\Downloads\T1 Measured CG_23-10-2025 (11-3-2026) updated 1 (1).xlsx"

SHEETS = {
    "Measured": "StrengthChk_TwistLock(Measured)",
    "Theory":   "StrengthChk_TwistLock(Theory)",
}
# The shelter weight cell (B7) and per-row cell columns (F=design, H=per-lock).
WEIGHT_CELL = "B7"
AXIS_ROWS = {"longitudinal": 6, "lateral": 7, "vertical": 8}
TOL = 1e-6


def main() -> int:
    path = os.environ.get("T1_MOBILITY_XLSX") or WB_DEFAULT
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"[ERROR] cannot open workbook: {e}")
        return 2

    n_pass = n_total = 0
    for variant, sheet_name in SHEETS.items():
        ws = wb[sheet_name]
        W = ws[WEIGHT_CELL].value
        report = analyze_twist_locks(float(W))
        print(f"=== {variant} (shelter {W} kg) — sheet {sheet_name!r} ===")
        print(f"  {'axis':<13}{'engine F':>16}{'stored F':>16}"
              f"{'engine/lock':>15}{'stored/lock':>15}")
        for axis, row in AXIS_ROWS.items():
            a = report.by_axis(axis)
            f_stored = ws.cell(row=row, column=6).value    # F
            per_stored = ws.cell(row=row, column=8).value  # H
            for label, eng, stored in (("F", a.design_force_N, f_stored),
                                       ("per", a.force_per_lock_N, per_stored)):
                n_total += 1
                if abs(eng - stored) <= TOL * max(1.0, abs(stored)):
                    n_pass += 1
                else:
                    print(f"  FAIL {axis} {label}: engine={eng} stored={stored}")
            print(f"  {axis:<13}{a.design_force_N:>16.4f}{f_stored:>16.4f}"
                  f"{a.force_per_lock_N:>15.4f}{per_stored:>15.4f}")
        print()

    print(f"TOTAL: {n_pass}/{n_total} pass")
    return 0 if n_pass == n_total else 1


if __name__ == "__main__":
    sys.exit(main())
